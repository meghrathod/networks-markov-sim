import os
from typing import List

import openpyxl


# Import the values from the environment.py file


class Result:
    total = 0

    def __init__(self, success: List[int], failure: List[int], timeOfExecution: int, throughput_avg: float):
        self.success = success
        self.failure = failure
        self.total = sum(success) + sum(failure)
        self.environment = ""
        self.timeOfExecution = timeOfExecution
        self.throughput_avg = throughput_avg

    def get_total_success(self):
        return sum(self.success)

    def get_success(self):
        return self.success

    def get_failure(self):
        return self.failure

    def get_total_failure(self):
        return sum(self.failure)

    def save_to_file(self, file_name: str, environment: str, time_to_trigger: int, hysteresis: int, velocity_min: float,
                     velocity_max: float):
        # Check if the file exists
        if not os.path.exists(file_name):
            print("File does not exist, creating new file")
            # Create a new workbook and add a header row
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.cell(row=1, column=1).value = "Environment"
            sheet.cell(row=1, column=2).value = "Time of execution"
            sheet.cell(row=1, column=3).value = "Time To Trigger"
            sheet.cell(row=1, column=4).value = "Hysteresis"
            sheet.cell(row=1, column=5).value = "Mean Velocity"
            sheet.cell(row=1, column=6).value = "Success [lte2lte]"
            sheet.cell(row=1, column=7).value = "Success [lte2nr]"
            sheet.cell(row=1, column=8).value = "Success [nr2lte]"
            sheet.cell(row=1, column=9).value = "Success [nr2nr]"
            sheet.cell(row=1, column=10).value = "Failure [lte2lte]"
            sheet.cell(row=1, column=11).value = "Failure [lte2nr]"
            sheet.cell(row=1, column=12).value = "Failure [nr2lte]"
            sheet.cell(row=1, column=13).value = "Failure [nr2nr]"
            sheet.cell(row=1, column=14).value = "Failed HOs"
            sheet.cell(row=1, column=15).value = "Successful HOs"
            sheet.cell(row=1, column=16).value = "Total HOs"
            sheet.cell(row=1, column=17).value = "Throughput Average"
        else:
            # Load the existing workbook
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active

        # Get the next available row on the sheet
        next_row = sheet.max_row + 1

        # Write the TTT and HYSTERESIS to the sheet
        sheet.cell(row=next_row, column=1).value = environment
        sheet.cell(row=next_row, column=2).value = self.timeOfExecution
        sheet.cell(row=next_row, column=3).value = time_to_trigger
        sheet.cell(row=next_row, column=4).value = hysteresis
        sheet.cell(row=next_row, column=5).value = (velocity_min + velocity_max) / 2

        # Write the elements of self.success to separate cells
        for i, element in enumerate(self.success):
            sheet.cell(row=next_row, column=i + 1 + 5).value = element

        # Write the elements of self.failure to separate cells
        for i, element in enumerate(self.failure):
            sheet.cell(row=next_row, column=i + 5 + 5).value = element

        # Write the total success and total failure to the sheet
        sheet.cell(row=next_row, column=14).value = (self.get_total_failure())
        sheet.cell(row=next_row, column=15).value = (self.get_total_success())
        sheet.cell(row=next_row, column=16).value = (self.get_total_ho())
        sheet.cell(row=next_row, column=17).value = (self.throughput_avg)

        # Save the workbook
        workbook.save(file_name)

    def get_total_ho(self):
        return self.total
